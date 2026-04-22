"""
Script de exportação do Decolog para Excel.
Coloca este arquivo na raiz do projeto (junto ao manage.py) e roda:
    python exportar_decolog.py
"""
import os
import sys
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
django.setup()

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from produtos.models import Produto
from estoque.models import Estoque


def estilo_cabecalho(cell):
    cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    cell.fill = PatternFill('solid', start_color='1a2535')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    borda = Side(style='thin', color='AAAAAA')
    cell.border = Border(left=borda, right=borda, top=borda, bottom=borda)


def estilo_dado(cell, par=False):
    cell.font = Font(name='Arial', size=10)
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    if par:
        cell.fill = PatternFill('solid', start_color='F0F4FA')
    borda = Side(style='thin', color='DDDDDD')
    cell.border = Border(left=borda, right=borda, top=borda, bottom=borda)


wb = Workbook()

# ── ABA 1: PRODUTOS ──────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Produtos'
ws1.row_dimensions[1].height = 30

cabecalhos = [
    'ID', 'Código', 'Nome', 'Categoria', 'Unidade de Medida',
    'Cor', 'Espessura (mm)', 'Largura (mm)', 'Comprimento (mm)',
    'Largura (cm)', 'Comprimento (cm)', 'Altura (cm)',
    'Diâmetro (cm)', 'Profundidade (cm)', 'Curvatura (cm)',
    'Descrição', 'Ativo',
]

larguras = [6, 12, 40, 18, 18, 12, 14, 12, 16, 12, 16, 12, 12, 16, 12, 40, 8]

for col, (cab, larg) in enumerate(zip(cabecalhos, larguras), 1):
    cell = ws1.cell(row=1, column=col, value=cab)
    estilo_cabecalho(cell)
    ws1.column_dimensions[get_column_letter(col)].width = larg

produtos = Produto.objects.all().order_by('categoria', 'nome')

for row_idx, p in enumerate(produtos, 2):
    par = row_idx % 2 == 0
    ws1.row_dimensions[row_idx].height = 18
    valores = [
        p.pk,
        p.codigo or '',
        p.nome,
        p.get_categoria_display(),
        p.get_unidade_medida_display(),
        p.get_cor_display() if p.cor else '',
        float(p.espessura_mm) if p.espessura_mm else '',
        float(p.largura_mm) if p.largura_mm else '',
        float(p.comprimento_mm) if p.comprimento_mm else '',
        float(p.largura_cm) if p.largura_cm else '',
        float(p.comprimento_cm) if p.comprimento_cm else '',
        float(p.altura_cm) if p.altura_cm else '',
        float(p.diametro_cm) if p.diametro_cm else '',
        float(p.profundidade_cm) if p.profundidade_cm else '',
        float(p.curvatura_cm) if p.curvatura_cm else '',
        p.descricao or '',
        'Sim' if p.ativo else 'Não',
    ]
    for col_idx, valor in enumerate(valores, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=valor)
        estilo_dado(cell, par)

ws1.freeze_panes = 'A2'

# ── ABA 2: ESTOQUE POR LOCAL ──────────────────────────────────────────────────
ws2 = wb.create_sheet('Estoque por Local')
ws2.row_dimensions[1].height = 30

cabecalhos2 = ['ID Produto', 'Produto', 'Categoria', 'Local', 'Tipo Local', 'Quantidade', 'Unidade']
larguras2 = [10, 40, 18, 25, 12, 12, 10]

for col, (cab, larg) in enumerate(zip(cabecalhos2, larguras2), 1):
    cell = ws2.cell(row=1, column=col, value=cab)
    estilo_cabecalho(cell)
    ws2.column_dimensions[get_column_letter(col)].width = larg

estoques = Estoque.objects.select_related('produto', 'local').order_by(
    'produto__categoria', 'produto__nome', 'local__nome'
)

for row_idx, e in enumerate(estoques, 2):
    par = row_idx % 2 == 0
    ws2.row_dimensions[row_idx].height = 18
    valores2 = [
        e.produto.pk,
        e.produto.nome,
        e.produto.get_categoria_display(),
        e.local.nome,
        e.local.get_tipo_display() if hasattr(e.local, 'get_tipo_display') else e.local.tipo,
        float(e.quantidade),
        e.produto.get_unidade_medida_display(),
    ]
    for col_idx, valor in enumerate(valores2, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=valor)
        estilo_dado(cell, par)

ws2.freeze_panes = 'A2'

# ── ABA 3: RESUMO ESTOQUE TOTAL ───────────────────────────────────────────────
ws3 = wb.create_sheet('Estoque Total por Produto')
ws3.row_dimensions[1].height = 30

cabecalhos3 = ['ID', 'Produto', 'Categoria', 'Unidade', 'Estoque Total']
larguras3 = [6, 40, 18, 12, 14]

for col, (cab, larg) in enumerate(zip(cabecalhos3, larguras3), 1):
    cell = ws3.cell(row=1, column=col, value=cab)
    estilo_cabecalho(cell)
    ws3.column_dimensions[get_column_letter(col)].width = larg

from django.db.models import Sum
resumo = (
    Estoque.objects
    .values('produto__pk', 'produto__nome', 'produto__categoria', 'produto__unidade_medida')
    .annotate(total=Sum('quantidade'))
    .order_by('produto__categoria', 'produto__nome')
)

for row_idx, r in enumerate(resumo, 2):
    par = row_idx % 2 == 0
    ws3.row_dimensions[row_idx].height = 18

    # Pega display da categoria e unidade
    p = Produto.objects.get(pk=r['produto__pk'])
    valores3 = [
        r['produto__pk'],
        r['produto__nome'],
        p.get_categoria_display(),
        p.get_unidade_medida_display(),
        float(r['total']),
    ]
    for col_idx, valor in enumerate(valores3, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=valor)
        estilo_dado(cell, par)

ws3.freeze_panes = 'A2'

# ── SALVA ─────────────────────────────────────────────────────────────────────
nome_arquivo = 'decolog_backup.xlsx'
wb.save(nome_arquivo)
print(f'✅ Exportado com sucesso: {nome_arquivo}')
print(f'   {ws1.max_row - 1} produtos')
print(f'   {ws2.max_row - 1} registros de estoque por local')
print(f'   {ws3.max_row - 1} produtos com estoque')